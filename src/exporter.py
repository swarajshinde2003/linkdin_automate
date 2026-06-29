"""Export HiringPost records to an Excel workbook."""
from __future__ import annotations

import io
from datetime import datetime

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.models import HiringPost


MAIN_COLS = ["role", "company", "location", "experience", "hr_mail", "post_link", "posted_at", "confidence", "matched_keywords", "source"]
REVIEW_COLS = MAIN_COLS + ["needs_review", "raw_text"]

# Human-friendly column header rename for the Excel output
_COL_RENAME = {"hr_mail": "contact_email"}


def to_excel_bytes(confirmed: list[HiringPost], review: list[HiringPost]) -> bytes:
    """Return an in-memory .xlsx file as bytes, ready for st.download_button."""
    buf = io.BytesIO()

    confirmed_rows = [p.to_row() for p in confirmed]
    review_rows = [p.to_review_row() for p in review]

    df_main = pd.DataFrame(confirmed_rows, columns=MAIN_COLS) if confirmed_rows else pd.DataFrame(columns=MAIN_COLS)
    df_review = pd.DataFrame(review_rows, columns=REVIEW_COLS) if review_rows else pd.DataFrame(columns=REVIEW_COLS)

    df_main = df_main.rename(columns=_COL_RENAME)
    df_review = df_review.rename(columns=_COL_RENAME)

    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df_main.to_excel(writer, sheet_name="Hiring Posts", index=False)
        df_review.to_excel(writer, sheet_name="Needs Review", index=False)

        _style_sheet(writer.sheets["Hiring Posts"], df_main)
        _style_sheet(writer.sheets["Needs Review"], df_review)

    return buf.getvalue()


def output_filename() -> str:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"linkedin_hiring_{ts}.xlsx"


# ── styling helpers ───────────────────────────────────────────────────────────────

def _style_sheet(ws, df: pd.DataFrame) -> None:
    header_fill = PatternFill("solid", fgColor="0A66C2")   # LinkedIn blue
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

        # Auto-width
        max_len = max(
            len(str(col_name)),
            *[len(str(v)) for v in df[col_name].values[:50]] if len(df) else [0],
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Wrap text in raw_text column if present
    if "raw_text" in df.columns:
        raw_col = list(df.columns).index("raw_text") + 1
        for row in ws.iter_rows(min_row=2, min_col=raw_col, max_col=raw_col):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)
